from hashlib import sha256
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Side

from core import (
    OperationRow,
    ProcessCard,
    generate_template_atomic,
    generate_template,
    group_operations,
    output_row_count,
    preview_rows,
)


SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


def qualified(namespace: str, name: str) -> str:
    return f"{{{namespace}}}{name}"


def make_shared_string_template(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "工艺路线编号",
            "工艺路线名称",
            "工序/工艺路线列表",
            "",
            "工序内容",
            "类型",
            "报工数配比",
            "是否锁定为最后一道工序",
            "工时/分钟",
        ]
    )
    sheet.append(["模板旧数据", "模板旧数据", "车", "10", "旧内容", "工序", 1])
    thin_black = Side(style="thin", color="FF000000")
    border = Border(
        left=thin_black,
        right=thin_black,
        top=thin_black,
        bottom=thin_black,
    )
    for cell in sheet[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in sheet[2][:7]:
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 30
    sheet["A1"].comment = Comment("保留的模板批注", "DocSwift")
    workbook.save(path)

    with ZipFile(path) as archive:
        infos = archive.infolist()
        parts = {info.filename: archive.read(info.filename) for info in infos}
    sheet_root = ElementTree.fromstring(parts["xl/worksheets/sheet1.xml"])
    strings: list[str] = []
    indexes: dict[str, int] = {}
    for cell in sheet_root.iter(qualified(SPREADSHEET_NS, "c")):
        if cell.attrib.get("t") != "inlineStr":
            continue
        inline = cell.find(qualified(SPREADSHEET_NS, "is"))
        if inline is None:
            continue
        text = "".join(
            node.text or "" for node in inline.iter(qualified(SPREADSHEET_NS, "t"))
        )
        index = indexes.get(text)
        if index is None:
            index = len(strings)
            indexes[text] = index
            strings.append(text)
        cell.remove(inline)
        cell.set("t", "s")
        value = ElementTree.SubElement(cell, qualified(SPREADSHEET_NS, "v"))
        value.text = str(index)
    parts["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
        sheet_root, encoding="utf-8", xml_declaration=True
    )

    shared_root = ElementTree.Element(
        qualified(SPREADSHEET_NS, "sst"),
        {"count": str(len(strings)), "uniqueCount": str(len(strings))},
    )
    for text in strings:
        item = ElementTree.SubElement(shared_root, qualified(SPREADSHEET_NS, "si"))
        text_node = ElementTree.SubElement(item, qualified(SPREADSHEET_NS, "t"))
        text_node.text = text
    parts["xl/sharedStrings.xml"] = ElementTree.tostring(
        shared_root, encoding="utf-8", xml_declaration=True
    )

    content_types = ElementTree.fromstring(parts["[Content_Types].xml"])
    ElementTree.SubElement(
        content_types,
        qualified(CONTENT_TYPES_NS, "Override"),
        {
            "PartName": "/xl/sharedStrings.xml",
            "ContentType": (
                "application/vnd.openxmlformats-officedocument.spreadsheetml."
                "sharedStrings+xml"
            ),
        },
    )
    parts["[Content_Types].xml"] = ElementTree.tostring(
        content_types, encoding="utf-8", xml_declaration=True
    )
    relationships = ElementTree.fromstring(parts["xl/_rels/workbook.xml.rels"])
    used_ids = {relationship.attrib["Id"] for relationship in relationships}
    relationship_number = 1
    while f"rId{relationship_number}" in used_ids:
        relationship_number += 1
    ElementTree.SubElement(
        relationships,
        qualified(PACKAGE_REL_NS, "Relationship"),
        {
            "Id": f"rId{relationship_number}",
            "Type": (
                "http://schemas.openxmlformats.org/officeDocument/2006/"
                "relationships/sharedStrings"
            ),
            "Target": "sharedStrings.xml",
        },
    )
    parts["xl/_rels/workbook.xml.rels"] = ElementTree.tostring(
        relationships, encoding="utf-8", xml_declaration=True
    )

    rewritten = path.with_name("shared-template.xlsx")
    with ZipFile(rewritten, "w", compression=ZIP_DEFLATED) as archive:
        written: set[str] = set()
        for info in infos:
            archive.writestr(info, parts[info.filename])
            written.add(info.filename)
        for name, data in parts.items():
            if name not in written:
                archive.writestr(name, data)
    rewritten.replace(path)


def workbook_parts(path: Path) -> dict[str, bytes]:
    with ZipFile(path) as archive:
        return {name: archive.read(name) for name in archive.namelist()}


class CoreRulesTest(unittest.TestCase):
    def make_card(self, operations: list[OperationRow]) -> ProcessCard:
        return ProcessCard(
            source_path=Path("sample.docx"),
            part_no="B.0001",
            part_name="测试件",
            operations=operations,
        )

    def assert_template_package_preserved(self, template: Path, output: Path) -> None:
        template_parts = workbook_parts(template)
        output_parts = workbook_parts(output)
        self.assertEqual(set(template_parts), set(output_parts))
        changed_parts = {"xl/sharedStrings.xml", "xl/worksheets/sheet1.xml"}
        for name, template_data in template_parts.items():
            if name in changed_parts:
                continue
            self.assertEqual(
                sha256(template_data).digest(),
                sha256(output_parts[name]).digest(),
                name,
            )

        sheet = ElementTree.fromstring(output_parts["xl/worksheets/sheet1.xml"])
        cells = list(sheet.iter(qualified(SPREADSHEET_NS, "c")))
        self.assertNotIn("inlineStr", {cell.attrib.get("t") for cell in cells})
        self.assertNotIn("str", {cell.attrib.get("t") for cell in cells})
        self.assertTrue(any(cell.attrib.get("t") == "s" for cell in cells))
        data_references = {
            cell.attrib.get("r")
            for cell in cells
            if int("".join(filter(str.isdigit, cell.attrib.get("r", "0"))) or "0") >= 2
        }
        self.assertFalse(any(reference.startswith(("H", "I")) for reference in data_references))

    def test_single_blank_continuation_is_grouped(self) -> None:
        card = self.make_card(
            [
                OperationRow("钻", "14", "第一步"),
                OperationRow("", "15", "第二步"),
                OperationRow("检验", "16", "检验"),
            ]
        )

        rows = preview_rows([card])

        self.assertEqual(2, len(rows))
        self.assertEqual(("14", "钻", "第一步\n第二步"), rows[0][2:])
        self.assertEqual(("16", "检验", "检验"), rows[1][2:])

    def test_multiple_blank_continuations_are_grouped(self) -> None:
        card = self.make_card(
            [
                OperationRow("车", "4", "步骤一"),
                OperationRow("", "5", "步骤二"),
                OperationRow("", "6", "步骤三"),
            ]
        )

        rows = preview_rows([card])

        self.assertEqual(1, len(rows))
        self.assertEqual("4", rows[0][2])
        self.assertEqual("步骤一\n步骤二\n步骤三", rows[0][4])

    def test_wait_weld_is_always_excluded(self) -> None:
        card = self.make_card(
            [
                OperationRow("清洗", "34", "清洗"),
                OperationRow("", "35", "待焊。"),
            ]
        )

        self.assertEqual(1, output_row_count([card]))
        self.assertNotIn("待焊", preview_rows([card])[0][4])

    def test_custom_exclusion_and_original_range(self) -> None:
        result = group_operations(
            [
                OperationRow("车", "4", "第一步"),
                OperationRow("", "5", "第二步"),
                OperationRow("外协", "6", "送外加工"),
                OperationRow("检验", "7", "检验"),
            ],
            ("待焊", "外协"),
        )

        self.assertEqual(2, len(result.operations))
        self.assertEqual("4～5", result.operations[0].original_range)
        self.assertEqual(["外协"], [operation.work_type for operation in result.excluded])

    def test_generation_keeps_operation_header_blank_and_formats_cells(self) -> None:
        card = self.make_card(
            [
                OperationRow("钻", "12", "第一步"),
                OperationRow("", "13", "第二步"),
                OperationRow("检验", "14", "检验"),
            ]
        )

        with TemporaryDirectory() as temporary_directory:
            directory = Path(temporary_directory)
            template = directory / "template.xlsx"
            output = directory / "output.xlsx"

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(
                [
                    "工艺路线编号",
                    "工艺路线名称",
                    "工序/工艺路线列表",
                    "",
                    "工序内容",
                    "类型",
                    "报工数配比",
                    "是否锁定为最后一道工序",
                    "工时/分钟",
                ]
            )
            sheet.append([None] * 9)
            workbook.save(template)

            generate_template([card], template, output, start_row=2)
            generated = load_workbook(output).active

            self.assertIn(generated["D1"].value, (None, ""))
            self.assertEqual("12", generated["D2"].value)
            self.assertEqual("第一步\n第二步", generated["E2"].value)
            self.assertEqual("14", generated["D3"].value)
            for row in range(1, 4):
                for column in range(1, 10):
                    cell = generated.cell(row, column)
                    self.assertEqual("center", cell.alignment.horizontal)
                    self.assertEqual("center", cell.alignment.vertical)
                    if row >= 2:
                        self.assertTrue(
                            all(
                                getattr(cell.border, side).style == "thin"
                                for side in ("left", "right", "top", "bottom")
                            )
                        )

    def test_atomic_export_replaces_route_without_duplicate_rows(self) -> None:
        original = self.make_card([OperationRow("车", "1", "旧内容")])
        updated = self.make_card([OperationRow("车", "1", "新内容")])

        with TemporaryDirectory() as temporary_directory:
            directory = Path(temporary_directory)
            template = directory / "template.xlsx"
            output = directory / "output.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(
                [
                    "工艺路线编号",
                    "工艺路线名称",
                    "工序/工艺路线列表",
                    "工序号",
                    "工序内容",
                    "类型",
                    "报工数配比",
                    "是否锁定为最后一道工序",
                    "工时/分钟",
                ]
            )
            sheet.append([None] * 9)
            workbook.save(template)

            generate_template_atomic([original], template, output)
            generate_template_atomic(
                [updated],
                template,
                output,
                replace_route_texts=[original.route_text],
            )

            generated = load_workbook(output).active
            values = [
                generated.cell(row, 5).value
                for row in range(2, generated.max_row + 1)
                if generated.cell(row, 1).value
            ]
            self.assertEqual(["新内容"], values)

    def test_atomic_export_keeps_operation_header_blank_and_preserves_numbers(self) -> None:
        card = self.make_card(
            [
                OperationRow("车", "10", "第一步"),
                OperationRow("", "11", "第二步"),
            ]
        )

        with TemporaryDirectory() as temporary_directory:
            directory = Path(temporary_directory)
            template = directory / "template.xlsx"
            output = directory / "output.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(
                [
                    "工艺路线编号",
                    "工艺路线名称",
                    "工序/工艺路线列表",
                    "",
                    "工序内容",
                    "类型",
                    "报工数配比",
                    "是否锁定为最后一道工序",
                    "工时/分钟",
                ]
            )
            sheet.append([None] * 9)
            workbook.save(template)

            generate_template_atomic([card], template, output)

            generated = load_workbook(output).active
            self.assertIn(generated["D1"].value, (None, ""))
            self.assertEqual("10", generated["D2"].value)
            self.assertEqual("第一步\n第二步", generated["E2"].value)

    def test_generation_writes_true_blanks_and_removes_stale_physical_rows(self) -> None:
        card = self.make_card(
            [
                OperationRow("车", "10", "第一步"),
                OperationRow("检验", "20", "第二步"),
            ]
        )

        with TemporaryDirectory() as temporary_directory:
            directory = Path(temporary_directory)
            template = directory / "template.xlsx"
            output = directory / "output.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(
                [
                    "工艺路线编号",
                    "工艺路线名称",
                    "工序/工艺路线列表",
                    "",
                    "工序内容",
                    "类型",
                    "报工数配比",
                    "是否锁定为最后一道工序",
                    "工时/分钟",
                ]
            )
            for row_number in range(2, 10):
                sheet.cell(row_number, 1).value = f"旧路线{row_number}"
            sheet.row_dimensions[2].height = 30
            sheet.row_dimensions[5].height = 45
            sheet.row_dimensions[9].height = 60
            workbook.save(template)

            generate_template([card], template, output, start_row=2)

            generated = load_workbook(output).active
            self.assertEqual(3, generated.max_row)
            self.assertEqual({2, 3}, set(generated.row_dimensions))
            for row_number in (2, 3):
                for column in (8, 9):
                    cell = generated.cell(row_number, column)
                    self.assertIsNone(cell.value)
                    self.assertNotEqual("inlineStr", cell.data_type)

            with ZipFile(output) as archive:
                worksheet_xml = ElementTree.fromstring(
                    archive.read("xl/worksheets/sheet1.xml")
                )
            namespace = {"main": worksheet_xml.tag.partition("}")[0].lstrip("{")}
            physical_rows = worksheet_xml.findall(".//main:sheetData/main:row", namespace)
            self.assertEqual([1, 2, 3], [int(row.attrib["r"]) for row in physical_rows])
            optional_references = {"H2", "I2", "H3", "I3"}
            optional_cells = [
                cell
                for cell in worksheet_xml.findall(".//main:c", namespace)
                if cell.attrib.get("r") in optional_references
            ]
            for cell in optional_cells:
                self.assertNotIn(cell.attrib.get("t"), {"s", "str", "inlineStr"})
                self.assertIsNone(cell.find("main:v", namespace))
                self.assertIsNone(cell.find("main:is", namespace))

    def test_atomic_replacement_removes_stale_tail_row_dimensions(self) -> None:
        retained = ProcessCard(
            source_path=Path("retained.docx"),
            part_no="B.0001",
            part_name="保留件",
            operations=[OperationRow("车", "10", "保留内容")],
        )
        original_tail = ProcessCard(
            source_path=Path("original.docx"),
            part_no="B.0002",
            part_name="替换件",
            operations=[
                OperationRow("车", "10", "旧内容一"),
                OperationRow("钻", "20", "旧内容二"),
                OperationRow("检验", "30", "旧内容三"),
            ],
        )
        replacement_tail = ProcessCard(
            source_path=Path("replacement.docx"),
            part_no="B.0002",
            part_name="替换件",
            operations=[OperationRow("车", "10", "新内容")],
        )

        with TemporaryDirectory() as temporary_directory:
            directory = Path(temporary_directory)
            template = directory / "template.xlsx"
            output = directory / "output.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(
                [
                    "工艺路线编号",
                    "工艺路线名称",
                    "工序/工艺路线列表",
                    "",
                    "工序内容",
                    "类型",
                    "报工数配比",
                    "是否锁定为最后一道工序",
                    "工时/分钟",
                ]
            )
            sheet.cell(2, 1).value = "模板旧数据"
            sheet.row_dimensions[2].height = 30
            workbook.save(template)

            generate_template_atomic([retained, original_tail], template, output)
            generate_template_atomic(
                [replacement_tail],
                template,
                output,
                replace_route_texts=[original_tail.route_text],
            )

            generated = load_workbook(output).active
            self.assertEqual(3, generated.max_row)
            self.assertEqual({2, 3}, set(generated.row_dimensions))
            with ZipFile(output) as archive:
                worksheet_xml = ElementTree.fromstring(
                    archive.read("xl/worksheets/sheet1.xml")
                )
            namespace = {"main": worksheet_xml.tag.partition("}")[0].lstrip("{")}
            physical_rows = worksheet_xml.findall(".//main:sheetData/main:row", namespace)
            self.assertEqual([1, 2, 3], [int(row.attrib["r"]) for row in physical_rows])

    def test_first_atomic_export_preserves_template_package_and_uses_shared_strings(self) -> None:
        card = self.make_card(
            [
                OperationRow("车", "10", "第一步"),
                OperationRow("检验", "20", "第二步"),
            ]
        )
        with TemporaryDirectory() as temporary_directory:
            directory = Path(temporary_directory)
            template = directory / "template.xlsx"
            output = directory / "output.xlsx"
            make_shared_string_template(template)

            generate_template_atomic([card], template, output)

            self.assert_template_package_preserved(template, output)
            generated = load_workbook(output)
            self.assertEqual(card.route_text, generated.active["A2"].value)
            self.assertEqual("第一步", generated.active["E2"].value)
            self.assertEqual("第二步", generated.active["E3"].value)
            self.assertEqual("保留的模板批注", generated.active["A1"].comment.text)

    def test_repeated_atomic_export_preserves_template_package_and_shared_strings(self) -> None:
        retained = ProcessCard(
            source_path=Path("retained.docx"),
            part_no="B.0001",
            part_name="保留件",
            operations=[OperationRow("车", "10", "保留内容")],
        )
        original = ProcessCard(
            source_path=Path("original.docx"),
            part_no="B.0002",
            part_name="替换件",
            operations=[OperationRow("钻", "20", "旧内容")],
        )
        updated = ProcessCard(
            source_path=Path("updated.docx"),
            part_no="B.0002",
            part_name="替换件",
            operations=[OperationRow("检验", "30", "新内容")],
        )
        with TemporaryDirectory() as temporary_directory:
            directory = Path(temporary_directory)
            template = directory / "template.xlsx"
            output = directory / "output.xlsx"
            make_shared_string_template(template)

            generate_template_atomic([retained, original], template, output)
            generate_template_atomic(
                [updated],
                template,
                output,
                replace_route_texts=[original.route_text],
            )

            self.assert_template_package_preserved(template, output)
            generated = load_workbook(output)
            routes = [generated.active.cell(row, 1).value for row in range(2, 4)]
            contents = [generated.active.cell(row, 5).value for row in range(2, 4)]
            self.assertEqual([retained.route_text, updated.route_text], routes)
            self.assertEqual(["保留内容", "新内容"], contents)
            self.assertEqual("保留的模板批注", generated.active["A1"].comment.text)


if __name__ == "__main__":
    unittest.main()
