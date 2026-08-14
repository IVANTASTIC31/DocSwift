from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from pathlib import Path
import os
import posixpath
import re
import shutil
import tempfile
from typing import Iterable, Sequence
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

from docx import Document
from openpyxl import load_workbook
from openpyxl.styles import Border, Side


HEADER_ALIASES = {
    "route_no": ("工艺路线编号",),
    "route_name": ("工艺路线名称",),
    "work_type": ("工序/工艺路线列表",),
    "content": ("工序内容",),
    "operation_no": ("工艺", "工序", "工序号"),
    "type": ("类型",),
    "ratio": ("报工数配比",),
    "locked": ("是否锁定为最后一道工序",),
    "work_minutes": ("工时/分钟",),
}

THIN_BLACK_BORDER = Border(
    left=Side(style="thin", color="FF000000"),
    right=Side(style="thin", color="FF000000"),
    top=Side(style="thin", color="FF000000"),
    bottom=Side(style="thin", color="FF000000"),
)

SPREADSHEET_XML_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
DOCUMENT_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
SHARED_STRINGS_PART = "xl/sharedStrings.xml"
SHARED_STRINGS_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"
)
SHARED_STRINGS_REL_TYPE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings"
)

ElementTree.register_namespace("", SPREADSHEET_XML_NS)
ElementTree.register_namespace("r", DOCUMENT_REL_NS)


@dataclass(frozen=True)
class OperationRow:
    work_type: str
    operation_no: str
    content: str


@dataclass(frozen=True)
class ProcessCard:
    source_path: Path
    part_no: str
    part_name: str
    operations: list[OperationRow]

    @property
    def route_text(self) -> str:
        return f"{self.part_no}{self.part_name}".strip()


@dataclass(frozen=True)
class FinalOperation:
    work_type: str
    operation_no: str
    original_range: str
    content: str


@dataclass(frozen=True)
class GroupingResult:
    operations: list[FinalOperation]
    excluded: list[OperationRow]


def group_operations(
    operations: Sequence[OperationRow],
    exclusion_terms: Sequence[str] = ("待焊",),
) -> GroupingResult:
    """Collapse continuation rows into the preceding named work type.

    Process cards commonly put the work type only on the first row of a
    multi-step operation.  Any following blank work-type row is treated as
    continuation detail. Operations matching an exclusion term are omitted
    before grouping. The result keeps the original numeric range for review,
    while Excel still receives the first operation number.
    """
    cleaned_terms = tuple(term.strip().casefold() for term in exclusion_terms if term.strip())
    included: list[OperationRow] = []
    excluded: list[OperationRow] = []
    for operation in operations:
        searchable = f"{operation.work_type}{operation.content}".casefold()
        if any(term in searchable for term in cleaned_terms):
            excluded.append(operation)
        else:
            included.append(operation)

    grouped: list[FinalOperation] = []
    index = 0
    while index < len(included):
        operation = included[index]
        if not operation.work_type:
            grouped.append(
                FinalOperation(
                    work_type=operation.work_type,
                    operation_no=operation.operation_no,
                    original_range=operation.operation_no,
                    content=operation.content,
                )
            )
            index += 1
            continue

        next_index = index + 1
        continuations: list[OperationRow] = []
        while next_index < len(included) and not included[next_index].work_type:
            continuations.append(included[next_index])
            next_index += 1

        if continuations:
            last_no = continuations[-1].operation_no
            grouped.append(
                FinalOperation(
                    work_type=operation.work_type,
                    operation_no=operation.operation_no,
                    original_range=(
                        operation.operation_no
                        if operation.operation_no == last_no
                        else f"{operation.operation_no}～{last_no}"
                    ),
                    content="\n".join(
                        item.content for item in (operation, *continuations) if item.content
                    ),
                )
            )
            index = next_index
        else:
            grouped.append(
                FinalOperation(
                    work_type=operation.work_type,
                    operation_no=operation.operation_no,
                    original_range=operation.operation_no,
                    content=operation.content,
                )
            )
            index += 1

    return GroupingResult(operations=grouped, excluded=excluded)


def _group_operations(operations: Sequence[OperationRow]) -> list[OperationRow]:
    result = group_operations(operations)
    return [
        OperationRow(
            work_type=operation.work_type,
            operation_no=operation.operation_no,
            content=operation.content,
        )
        for operation in result.operations
    ]


def output_row_count(cards: Sequence[ProcessCard]) -> int:
    return sum(len(_group_operations(card.operations)) for card in cards)


def _content_row_height(content: str, base_height: float | None) -> float | None:
    line_count = content.count("\n") + 1
    if line_count <= 1:
        return base_height
    # Short groups tend to wrap within the wide content column, while long
    # groups are more compact.  These values reproduce the supplied reference
    # workbook (6 steps -> 177 pt; 16 steps -> 242 pt).
    estimated_height = line_count * 28.5 + 6 if line_count <= 8 else line_count * 15 + 2
    return max(base_height or 0, estimated_height)


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u3000", " ")
    text = text.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    return text.strip()


def _key_text(value: object) -> str:
    return re.sub(r"[\s|:：]+", "", _clean_text(value))


def _dedupe_repeated_cells(values: Sequence[str]) -> list[str]:
    result: list[str] = []
    for value in values:
        if not result or result[-1] != value:
            result.append(value)
    return result


def _cell_texts(row) -> list[str]:
    return [_clean_text(cell.text) for cell in row.cells]


def _build_column_map(worksheet, header_row: int = 1) -> dict[str, int]:
    alias_lookup: dict[str, str] = {}
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            alias_lookup[_key_text(alias)] = field

    column_map: dict[str, int] = {}
    for col in range(1, worksheet.max_column + 1):
        key = _key_text(worksheet.cell(header_row, col).value)
        field = alias_lookup.get(key)
        if field and field not in column_map:
            column_map[field] = col

    fallback = {
        "route_no": 1,
        "route_name": 2,
        "work_type": 3,
        "operation_no": 4,
        "content": 5,
        "type": 6,
        "ratio": 7,
        "locked": 8,
        "work_minutes": 9,
    }
    for field, col in fallback.items():
        column_map.setdefault(field, col)

    return column_map


def _find_next_output_row(worksheet, start_row: int, column_map: dict[str, int]) -> int:
    content_columns = sorted(
        {
            column_map[field]
            for field in ("route_no", "route_name", "work_type", "content", "operation_no")
            if field in column_map
        }
    )
    if not content_columns:
        content_columns = list(range(1, worksheet.max_column + 1))

    for row_index in range(worksheet.max_row, start_row - 1, -1):
        if any(_clean_text(worksheet.cell(row_index, col).value) for col in content_columns):
            return row_index + 1
    return start_row


def _find_value_below(table, labels: Iterable[str]) -> str:
    wanted = set(labels)
    for row_index, row in enumerate(table.rows[:-1]):
        keys = [_key_text(cell.text) for cell in row.cells]
        for col_index, key in enumerate(keys):
            if key in wanted:
                for next_index in range(row_index + 1, min(row_index + 4, len(table.rows))):
                    value = _clean_text(table.cell(next_index, col_index).text)
                    if value and _key_text(value) not in wanted:
                        return value
    return ""


def _find_header(table) -> tuple[int, int, int, int]:
    best: tuple[int, int, int, int] | None = None
    for row_index, row in enumerate(table.rows):
        keys = [_key_text(cell.text) for cell in row.cells]
        try:
            work_col = keys.index("工种")
            no_col = keys.index("工序")
            content_col = keys.index("工序内容")
        except ValueError:
            continue
        best = (row_index, work_col, no_col, content_col)

    if best is None:
        raise ValueError("未找到包含“工种 / 工序 / 工序内容”的表头。")

    header_row, work_col, no_col, content_col = best
    return header_row, work_col, no_col, content_col


def _looks_like_operation_no(value: str) -> bool:
    return bool(re.fullmatch(r"\d+(?:[.\-]\d+)?", value.strip()))


def parse_process_card(path: str | Path) -> ProcessCard:
    source_path = Path(path)
    document = Document(source_path)

    part_no = ""
    part_name = ""
    operations: list[OperationRow] = []

    for table in document.tables:
        if not part_name:
            part_name = _find_value_below(table, {"零件名称"})
        if not part_no:
            part_no = _find_value_below(table, {"零件图号"})

        try:
            header_row, work_col, no_col, content_col = _find_header(table)
        except ValueError:
            continue

        for row in table.rows[header_row + 1 :]:
            values = _cell_texts(row)
            if max(work_col, no_col, content_col) >= len(values):
                continue

            work_type = values[work_col]
            operation_no = values[no_col]
            content = values[content_col]

            if not _looks_like_operation_no(operation_no):
                continue
            if not content:
                continue

            operations.append(
                OperationRow(
                    work_type=work_type,
                    operation_no=operation_no,
                    content=content,
                )
            )

    if not part_no:
        raise ValueError("未能从工艺卡中识别“零件图号”。")
    if not part_name:
        raise ValueError("未能从工艺卡中识别“零件名称”。")
    if not operations:
        raise ValueError("未能从工艺卡中识别任何工序明细。")

    return ProcessCard(
        source_path=source_path,
        part_no=part_no,
        part_name=part_name,
        operations=operations,
    )


def parse_many_cards(paths: Iterable[str | Path]) -> list[ProcessCard]:
    cards = [parse_process_card(path) for path in paths]
    if not cards:
        raise ValueError("请至少选择一个工艺卡 Word 文件。")
    return cards


def _copy_row_style(ws, template_row: int, target_row: int, max_col: int) -> None:
    if template_row <= 0:
        return

    ws.row_dimensions[target_row].height = ws.row_dimensions[template_row].height
    for col in range(1, max_col + 1):
        source = ws.cell(template_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)
        if source.protection:
            target.protection = copy(source.protection)


def _generate_template_openpyxl(
    cards: Sequence[ProcessCard],
    template_path: str | Path,
    output_path: str | Path,
    start_row: int = 6,
    append: bool = False,
) -> Path:
    if start_row < 2:
        raise ValueError("起始行不能小于 2。")
    if not cards:
        raise ValueError("没有可写入的工艺卡数据。")

    template_path = Path(template_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    append_to_existing = append and output_path.exists()
    workbook = load_workbook(output_path if append_to_existing else template_path)
    worksheet = workbook.active
    column_map = _build_column_map(worksheet)
    max_col = max(worksheet.max_column, max(column_map.values(), default=8), 8)
    output_max_col = max(column_map.values(), default=9)
    operation_no_col = column_map.get("operation_no")
    if operation_no_col:
        worksheet.cell(1, operation_no_col).value = None
    for col in range(1, output_max_col + 1):
        header_cell = worksheet.cell(1, col)
        alignment = copy(header_cell.alignment)
        alignment.horizontal = "center"
        alignment.vertical = "center"
        alignment.wrap_text = True
        header_cell.alignment = alignment

    row_index = _find_next_output_row(worksheet, start_row, column_map) if append_to_existing else start_row
    style_source_row = (
        max(start_row, row_index - 1)
        if append_to_existing and row_index > start_row
        else start_row if worksheet.max_row >= start_row else max(1, worksheet.max_row)
    )
    template_height = worksheet.row_dimensions[style_source_row].height
    template_styles = []
    for col in range(1, max_col + 1):
        cell = worksheet.cell(style_source_row, col)
        template_styles.append(
            {
                "style": copy(cell._style) if cell.has_style else None,
                "number_format": cell.number_format,
                "alignment": copy(cell.alignment),
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
                "protection": copy(cell.protection),
            }
        )

    if not append_to_existing and worksheet.max_row >= start_row:
        worksheet.delete_rows(start_row, worksheet.max_row - start_row + 1)
        for row_number in tuple(worksheet.row_dimensions):
            if row_number >= start_row:
                del worksheet.row_dimensions[row_number]

    for card in cards:
        for operation in _group_operations(card.operations):
            for col, style in enumerate(template_styles, start=1):
                target = worksheet.cell(row_index, col)
                if style["style"] is not None:
                    target._style = copy(style["style"])
                target.number_format = style["number_format"]
                target.alignment = copy(style["alignment"])
                target.font = copy(style["font"])
                target.fill = copy(style["fill"])
                target.border = copy(style["border"])
                target.protection = copy(style["protection"])
            worksheet.row_dimensions[row_index].height = _content_row_height(
                operation.content, template_height
            )

            values = {
                "route_no": card.route_text,
                "route_name": card.route_text,
                "work_type": operation.work_type,
                "content": operation.content,
                "operation_no": operation.operation_no,
                "type": "工序",
                "ratio": 1,
                "locked": None,
                "work_minutes": None,
            }
            for field, value in values.items():
                col = column_map.get(field)
                if col:
                    worksheet.cell(row_index, col).value = value
            for col in range(1, output_max_col + 1):
                target_cell = worksheet.cell(row_index, col)
                target_cell.border = copy(THIN_BLACK_BORDER)
                alignment = copy(target_cell.alignment)
                alignment.horizontal = "center"
                alignment.vertical = "center"
                alignment.wrap_text = True
                target_cell.alignment = alignment
            row_index += 1

    workbook.save(output_path)
    return output_path


def _qualified(namespace: str, name: str) -> str:
    return f"{{{namespace}}}{name}"


def _xml_bytes(root: ElementTree.Element) -> bytes:
    return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)


def _active_sheet_part(parts: dict[str, bytes]) -> str:
    workbook = ElementTree.fromstring(parts["xl/workbook.xml"])
    sheets = workbook.find(_qualified(SPREADSHEET_XML_NS, "sheets"))
    if sheets is None or not list(sheets):
        raise ValueError("Excel 工作簿没有可用工作表。")
    workbook_view = workbook.find(
        f"{_qualified(SPREADSHEET_XML_NS, 'bookViews')}/"
        f"{_qualified(SPREADSHEET_XML_NS, 'workbookView')}"
    )
    active_index = (
        int(workbook_view.attrib.get("activeTab", "0"))
        if workbook_view is not None
        else 0
    )
    sheet_nodes = list(sheets)
    active_index = min(max(active_index, 0), len(sheet_nodes) - 1)
    relationship_id = sheet_nodes[active_index].attrib.get(
        _qualified(DOCUMENT_REL_NS, "id")
    )
    if not relationship_id:
        raise ValueError("Excel 活动工作表缺少关系标识。")

    relationships = ElementTree.fromstring(parts["xl/_rels/workbook.xml.rels"])
    for relationship in relationships:
        if relationship.attrib.get("Id") != relationship_id:
            continue
        target = relationship.attrib.get("Target", "")
        if target.startswith("/"):
            return target.lstrip("/")
        return posixpath.normpath(posixpath.join("xl", target))
    raise ValueError("Excel 活动工作表关系不存在。")


def _shared_string_text(item: ElementTree.Element) -> str:
    return "".join(
        node.text or ""
        for node in item.iter(_qualified(SPREADSHEET_XML_NS, "t"))
    )


def _shared_string_items(data: bytes | None) -> tuple[ElementTree.Element, list[str]]:
    if data:
        root = ElementTree.fromstring(data)
    else:
        root = ElementTree.Element(_qualified(SPREADSHEET_XML_NS, "sst"))
    return root, [_shared_string_text(item) for item in root]


def _append_shared_string(root: ElementTree.Element, text: str) -> None:
    item = ElementTree.SubElement(root, _qualified(SPREADSHEET_XML_NS, "si"))
    text_node = ElementTree.SubElement(item, _qualified(SPREADSHEET_XML_NS, "t"))
    if text != text.strip():
        text_node.set(_qualified("http://www.w3.org/XML/1998/namespace", "space"), "preserve")
    text_node.text = text


def _convert_sheet_strings(
    sheet: ElementTree.Element,
    base_shared_strings: bytes | None,
    staging_shared_strings: bytes | None,
) -> ElementTree.Element:
    shared_root, shared_values = _shared_string_items(base_shared_strings)
    _, staging_values = _shared_string_items(staging_shared_strings)
    shared_indexes: dict[str, int] = {}
    for index, value in enumerate(shared_values):
        shared_indexes.setdefault(value, index)

    for cell in sheet.iter(_qualified(SPREADSHEET_XML_NS, "c")):
        data_type = cell.attrib.get("t")
        value: str | None = None
        inline_value = cell.find(_qualified(SPREADSHEET_XML_NS, "is"))
        value_node = cell.find(_qualified(SPREADSHEET_XML_NS, "v"))
        if data_type == "inlineStr" and inline_value is not None:
            value = _shared_string_text(inline_value)
        elif data_type == "s" and value_node is not None:
            try:
                value = staging_values[int(value_node.text or "0")]
            except (IndexError, ValueError):
                raise ValueError("Excel 临时文件包含无效共享字符串索引。") from None
        elif (
            data_type == "str"
            and value_node is not None
            and cell.find(_qualified(SPREADSHEET_XML_NS, "f")) is None
        ):
            value = value_node.text or ""
        if value is None:
            continue

        shared_index = shared_indexes.get(value)
        if shared_index is None:
            shared_index = len(shared_values)
            shared_values.append(value)
            shared_indexes[value] = shared_index
            _append_shared_string(shared_root, value)
        if inline_value is not None:
            cell.remove(inline_value)
        if value_node is not None:
            cell.remove(value_node)
        cell.set("t", "s")
        new_value_node = ElementTree.SubElement(
            cell, _qualified(SPREADSHEET_XML_NS, "v")
        )
        new_value_node.text = str(shared_index)

    shared_root.set("uniqueCount", str(len(shared_values)))
    return shared_root


def _replace_sheet_data(
    template_sheet: bytes,
    staging_sheet: ElementTree.Element,
) -> bytes:
    template_root = ElementTree.fromstring(template_sheet)
    template_data = template_root.find(_qualified(SPREADSHEET_XML_NS, "sheetData"))
    staging_data = staging_sheet.find(_qualified(SPREADSHEET_XML_NS, "sheetData"))
    if template_data is None or staging_data is None:
        raise ValueError("Excel 工作表缺少 sheetData。")
    data_index = list(template_root).index(template_data)
    template_root.remove(template_data)
    template_root.insert(data_index, staging_data)

    template_dimension = template_root.find(_qualified(SPREADSHEET_XML_NS, "dimension"))
    staging_dimension = staging_sheet.find(_qualified(SPREADSHEET_XML_NS, "dimension"))
    if template_dimension is not None and staging_dimension is not None:
        template_dimension.set("ref", staging_dimension.attrib.get("ref", "A1"))
    template_filter = template_root.find(_qualified(SPREADSHEET_XML_NS, "autoFilter"))
    staging_filter = staging_sheet.find(_qualified(SPREADSHEET_XML_NS, "autoFilter"))
    if template_filter is not None and staging_filter is not None:
        template_filter.set("ref", staging_filter.attrib.get("ref", "A1"))
    return _xml_bytes(template_root)


def _ensure_shared_strings_package(parts: dict[str, bytes]) -> None:
    content_types = ElementTree.fromstring(parts["[Content_Types].xml"])
    override_tag = _qualified(CONTENT_TYPES_NS, "Override")
    if not any(
        node.attrib.get("PartName") == "/xl/sharedStrings.xml"
        for node in content_types.findall(override_tag)
    ):
        ElementTree.SubElement(
            content_types,
            override_tag,
            {
                "PartName": "/xl/sharedStrings.xml",
                "ContentType": SHARED_STRINGS_CONTENT_TYPE,
            },
        )
        parts["[Content_Types].xml"] = _xml_bytes(content_types)

    relationships = ElementTree.fromstring(parts["xl/_rels/workbook.xml.rels"])
    relationship_tag = _qualified(PACKAGE_REL_NS, "Relationship")
    if not any(
        node.attrib.get("Type") == SHARED_STRINGS_REL_TYPE
        for node in relationships.findall(relationship_tag)
    ):
        used_ids = {node.attrib.get("Id") for node in relationships.findall(relationship_tag)}
        relationship_number = 1
        while f"rId{relationship_number}" in used_ids:
            relationship_number += 1
        ElementTree.SubElement(
            relationships,
            relationship_tag,
            {
                "Id": f"rId{relationship_number}",
                "Type": SHARED_STRINGS_REL_TYPE,
                "Target": "sharedStrings.xml",
            },
        )
        parts["xl/_rels/workbook.xml.rels"] = _xml_bytes(relationships)


def _cell_style_count(styles: bytes) -> int:
    root = ElementTree.fromstring(styles)
    cell_styles = root.find(_qualified(SPREADSHEET_XML_NS, "cellXfs"))
    return len(cell_styles) if cell_styles is not None else 0


def _column_number(cell_reference: str) -> int:
    letters = "".join(character for character in cell_reference if character.isalpha())
    result = 0
    for character in letters.upper():
        result = result * 26 + ord(character) - ord("A") + 1
    return result


def _template_export_style_map(
    template_sheet: bytes,
    template_styles: bytes,
    start_row: int,
) -> dict[int, int] | None:
    sheet_root = ElementTree.fromstring(template_sheet)
    source_row = next(
        (
            row
            for row in sheet_root.iter(_qualified(SPREADSHEET_XML_NS, "row"))
            if int(row.attrib.get("r", "0")) == start_row
        ),
        None,
    )
    if source_row is None:
        return None
    style_map = {
        _column_number(cell.attrib.get("r", "")): int(cell.attrib.get("s", "0"))
        for cell in source_row.findall(_qualified(SPREADSHEET_XML_NS, "c"))
        if cell.attrib.get("r")
    }
    if any(column not in style_map for column in range(1, 8)):
        return None

    styles_root = ElementTree.fromstring(template_styles)
    cell_styles = styles_root.find(_qualified(SPREADSHEET_XML_NS, "cellXfs"))
    borders = styles_root.find(_qualified(SPREADSHEET_XML_NS, "borders"))
    if cell_styles is None or borders is None:
        return None
    for column in range(1, 8):
        style_id = style_map[column]
        if style_id >= len(cell_styles):
            return None
        style = cell_styles[style_id]
        alignment = style.find(_qualified(SPREADSHEET_XML_NS, "alignment"))
        if alignment is None or any(
            (
                alignment.attrib.get("horizontal") != "center",
                alignment.attrib.get("vertical") != "center",
                alignment.attrib.get("wrapText") not in {"1", "true"},
            )
        ):
            return None
        border_id = int(style.attrib.get("borderId", "0"))
        if border_id >= len(borders):
            return None
        border = borders[border_id]
        if any(
            (border.find(_qualified(SPREADSHEET_XML_NS, side)) is None)
            or border.find(_qualified(SPREADSHEET_XML_NS, side)).attrib.get("style")
            != "thin"
            for side in ("left", "right", "top", "bottom")
        ):
            return None
    return style_map


def _reuse_template_data_styles(
    staging_sheet: ElementTree.Element,
    template_sheet: bytes,
    style_map: dict[int, int],
    start_row: int,
) -> None:
    template_root = ElementTree.fromstring(template_sheet)
    template_rows = {
        int(row.attrib.get("r", "0")): row
        for row in template_root.iter(_qualified(SPREADSHEET_XML_NS, "row"))
    }
    source_row = template_rows[start_row]
    for row in staging_sheet.iter(_qualified(SPREADSHEET_XML_NS, "row")):
        row_number = int(row.attrib.get("r", "0"))
        if row_number < start_row:
            template_row = template_rows.get(row_number)
            if template_row is not None:
                row.attrib.clear()
                row.attrib.update(template_row.attrib)
            continue
        for attribute in ("s", "customFormat", "spans"):
            row.attrib.pop(attribute, None)
            if attribute in source_row.attrib:
                row.set(attribute, source_row.attrib[attribute])
        for cell in list(row.findall(_qualified(SPREADSHEET_XML_NS, "c"))):
            column = _column_number(cell.attrib.get("r", ""))
            if column in style_map:
                cell.set("s", str(style_map[column]))
            elif column in (8, 9) and not list(cell):
                row.remove(cell)


def _shared_string_reference_count(parts: dict[str, bytes]) -> int:
    count = 0
    for name, data in parts.items():
        if not name.startswith("xl/worksheets/") or not name.endswith(".xml"):
            continue
        root = ElementTree.fromstring(data)
        count += sum(
            1
            for cell in root.iter(_qualified(SPREADSHEET_XML_NS, "c"))
            if cell.attrib.get("t") == "s"
        )
    return count


def _repack_from_template(
    template_path: Path,
    staging_path: Path,
    output_path: Path,
    start_row: int,
) -> None:
    with ZipFile(template_path) as template_archive:
        template_infos = template_archive.infolist()
        parts = {info.filename: template_archive.read(info.filename) for info in template_infos}
    with ZipFile(staging_path) as staging_archive:
        staging_parts = {
            info.filename: staging_archive.read(info.filename)
            for info in staging_archive.infolist()
        }

    template_sheet_part = _active_sheet_part(parts)
    staging_sheet_part = _active_sheet_part(staging_parts)
    staging_sheet = ElementTree.fromstring(staging_parts[staging_sheet_part])
    template_style_map = _template_export_style_map(
        parts[template_sheet_part], parts["xl/styles.xml"], start_row
    )
    if template_style_map is not None:
        _reuse_template_data_styles(
            staging_sheet,
            parts[template_sheet_part],
            template_style_map,
            start_row,
        )
    shared_root = _convert_sheet_strings(
        staging_sheet,
        parts.get(SHARED_STRINGS_PART),
        staging_parts.get(SHARED_STRINGS_PART),
    )
    parts[template_sheet_part] = _replace_sheet_data(
        parts[template_sheet_part], staging_sheet
    )
    parts[SHARED_STRINGS_PART] = _xml_bytes(shared_root)
    _ensure_shared_strings_package(parts)

    referenced_styles = [
        int(cell.attrib["s"])
        for cell in staging_sheet.iter(_qualified(SPREADSHEET_XML_NS, "c"))
        if "s" in cell.attrib
    ]
    if referenced_styles and max(referenced_styles) >= _cell_style_count(parts["xl/styles.xml"]):
        parts["xl/styles.xml"] = staging_parts["xl/styles.xml"]

    shared_root.set("count", str(_shared_string_reference_count(parts)))
    parts[SHARED_STRINGS_PART] = _xml_bytes(shared_root)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    file_descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{output_path.stem}.package.",
        suffix=".tmp.xlsx",
        dir=output_path.parent,
    )
    os.close(file_descriptor)
    temporary_path = Path(temporary_name)
    try:
        with ZipFile(temporary_path, "w", compression=ZIP_DEFLATED) as output_archive:
            written: set[str] = set()
            for info in template_infos:
                output_archive.writestr(info, parts[info.filename])
                written.add(info.filename)
            for name, data in parts.items():
                if name not in written:
                    output_archive.writestr(name, data)
        os.replace(temporary_path, output_path)
    except Exception:
        temporary_path.unlink(missing_ok=True)
        raise


def generate_template(
    cards: Sequence[ProcessCard],
    template_path: str | Path,
    output_path: str | Path,
    start_row: int = 6,
    append: bool = False,
) -> Path:
    template_path = Path(template_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    append_to_existing = append and output_path.exists()
    file_descriptor, staging_name = tempfile.mkstemp(
        prefix=f".{output_path.stem}.openpyxl.",
        suffix=".tmp.xlsx",
        dir=output_path.parent,
    )
    os.close(file_descriptor)
    staging_path = Path(staging_name)
    try:
        if append_to_existing:
            shutil.copy2(output_path, staging_path)
        else:
            staging_path.unlink(missing_ok=True)
        _generate_template_openpyxl(
            cards,
            template_path,
            staging_path,
            start_row=start_row,
            append=append_to_existing,
        )
        _repack_from_template(template_path, staging_path, output_path, start_row)
    finally:
        staging_path.unlink(missing_ok=True)
    return output_path


def remove_routes_from_workbook(
    workbook_path: str | Path,
    route_texts: Iterable[str],
    *,
    start_row: int = 2,
) -> int:
    """Remove previously exported routes from an existing workbook.

    Route text is unique within one target workbook by business rule. Rows are
    removed from bottom to top so unrelated routes and their relative order are
    preserved.
    """
    wanted = {_clean_text(value).casefold() for value in route_texts if _clean_text(value)}
    if not wanted:
        return 0
    workbook_path = Path(workbook_path)
    workbook = load_workbook(workbook_path)
    worksheet = workbook.active
    column_map = _build_column_map(worksheet)
    route_column = column_map["route_no"]
    rows_to_delete = [
        row_index
        for row_index in range(start_row, worksheet.max_row + 1)
        if _clean_text(worksheet.cell(row_index, route_column).value).casefold() in wanted
    ]
    for row_index in reversed(rows_to_delete):
        worksheet.delete_rows(row_index, 1)
    for row_number in tuple(worksheet.row_dimensions):
        if row_number > worksheet.max_row:
            del worksheet.row_dimensions[row_number]
    workbook.save(workbook_path)
    return len(rows_to_delete)


def generate_template_atomic(
    cards: Sequence[ProcessCard],
    template_path: str | Path,
    output_path: str | Path,
    *,
    start_row: int = 2,
    replace_route_texts: Iterable[str] = (),
) -> Path:
    """Write all cards as one transaction and leave the old output untouched on failure."""
    if not cards:
        raise ValueError("没有已确认且未导出的工艺卡。")

    template_path = Path(template_path)
    output_path = Path(output_path)
    if not template_path.exists():
        raise FileNotFoundError(f"Excel 模板不存在：{template_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    file_descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{output_path.stem}.",
        suffix=".tmp.xlsx",
        dir=output_path.parent,
    )
    os.close(file_descriptor)
    temporary_path = Path(temporary_name)
    try:
        if output_path.exists():
            shutil.copy2(output_path, temporary_path)
            remove_routes_from_workbook(
                temporary_path,
                replace_route_texts,
                start_row=start_row,
            )
            generate_template(
                cards,
                template_path,
                temporary_path,
                start_row=start_row,
                append=True,
            )
        else:
            generate_template(
                cards,
                template_path,
                temporary_path,
                start_row=start_row,
                append=False,
            )
        os.replace(temporary_path, output_path)
    except Exception:
        temporary_path.unlink(missing_ok=True)
        raise
    return output_path


def preview_rows(cards: Sequence[ProcessCard]) -> list[tuple[str, str, str, str, str]]:
    rows: list[tuple[str, str, str, str, str]] = []
    for card in cards:
        for operation in _group_operations(card.operations):
            rows.append(
                (
                    card.source_path.name,
                    card.route_text,
                    operation.operation_no,
                    operation.work_type,
                    operation.content,
                )
            )
    return rows
