from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from docx import Document
from openpyxl import Workbook, load_workbook

from domain import CardStatus
from project_store import ProjectStore
from services import export_confirmed_cards, recognize_docx


def make_process_card(path: Path) -> None:
    document = Document()
    table = document.add_table(rows=7, cols=3)
    table.cell(0, 0).text = "零件名称"
    table.cell(0, 1).text = "零件图号"
    table.cell(1, 0).text = "测试件"
    table.cell(1, 1).text = "B.0001"
    table.cell(2, 0).text = "工种"
    table.cell(2, 1).text = "工序"
    table.cell(2, 2).text = "工序内容"
    rows = (
        ("车", "4", "第一步"),
        ("", "5", "第二步"),
        ("外协", "6", "送外加工"),
        ("检验", "7", "检验"),
    )
    for row_index, values in enumerate(rows, start=3):
        for column, value in enumerate(values):
            table.cell(row_index, column).text = value
    document.save(path)


def make_template(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "工艺路线编号",
            "工艺路线名称",
            "工序/工艺路线列表",
            "工序号",
            "工序内容",
            "类型",
            "报工数配比",
            "是否锁定为最后一道工序",
            "工时/分钟",
        ]
    )
    sheet.append([None] * 9)
    workbook.save(path)


class ServicesTest(unittest.TestCase):
    def test_recognition_applies_custom_exclusion_and_merges(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            source = Path(temporary_directory) / "card.docx"
            make_process_card(source)

            payload = recognize_docx(source, ("待焊", "外协"))

            self.assertEqual("B.0001", payload.route_no)
            self.assertEqual("测试件", payload.route_name)
            self.assertEqual(["4", "7"], [row["operation_no"] for row in payload.operations])
            self.assertEqual("4～5", payload.operations[0]["original_range"])
            self.assertEqual(1, len(payload.excluded_snapshot))

    def test_confirmed_cards_export_together_and_status_is_saved(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            directory = Path(temporary_directory)
            database = directory / "docswift.db"
            template = directory / "template.xlsx"
            output = directory / "output.xlsx"
            make_template(template)

            with ProjectStore(database) as store:
                task = store.get_or_create_active_task()
                store.update_task(
                    task.id,
                    template_path=template,
                    output_path=output,
                )
                for index in (1, 2):
                    source = directory / f"B.{index}.docx"
                    source.write_bytes(f"source-{index}".encode())
                    card = store.add_cards(task.id, [source])[0]
                    store.replace_recognition(
                        card.id,
                        route_no=f"B.{index}",
                        route_name=f"测试件{index}",
                        operations=[
                            {
                                "operation_no": "1",
                                "original_range": "1",
                                "work_type": "检验",
                                "content": f"检验{index}",
                            }
                        ],
                        original_snapshot=[],
                        excluded_snapshot=[],
                    )
                    store.confirm_card(card.id)

            generated, card_ids, row_count = export_confirmed_cards(database, task.id)

            self.assertEqual(output, generated)
            self.assertEqual(2, len(card_ids))
            self.assertEqual(2, row_count)
            workbook = load_workbook(output)
            self.assertEqual(["B.1测试件1", "B.2测试件2"], [workbook.active["A2"].value, workbook.active["A3"].value])
            with ProjectStore(database) as store:
                self.assertTrue(
                    all(
                        card.status == CardStatus.EXPORTED
                        for card in store.list_cards(task.id)
                    )
                )


if __name__ == "__main__":
    unittest.main()
